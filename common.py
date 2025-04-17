from selenium.webdriver.edge.service import Service
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from collections import namedtuple
from openpyxl import load_workbook
from selenium import webdriver
from tkinter import messagebox
from logger import logger
import tkinter as tk
import pandas as pd
import win32print
import win32gui
import win32con
import pyautogui
import openpyxl
import config
import random
import ctypes
import time
import sys
import re
import os

position = namedtuple('position',
                      ['left', 'top', 'width', 'height'])

corpInfo = namedtuple('corpInfo',
                      ['serial', 'name', 'type', 'taxid', 'loginid', 'taxpwd', 'ptaxpwd'])

deviceInfo = namedtuple('DeviceInfo',
                        ['available', 'logic_w', 'logic_h', 'phys_w', 'phys_h', 'scale_factor'])


def run_as_admin():
    def is_admin():
        try:
            return ctypes.windll.shell32.IsUserAnAdmin()
        except:
            return False

    if not is_admin():
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 0)
        sys.exit()
    return True


def stop():
    pop_error_window('程序暂停，点击确定以继续', '程序暂停')


def set_window_position_size(window_title, x, y, w, h):
    try:
        # 查找窗口句柄
        hwnd = win32gui.FindWindow(None, window_title)
        if hwnd == 0:
            raise Exception(f"未找到标题为 '{window_title}' 的窗口")

        # 获取窗口的显示状态
        placement = win32gui.GetWindowPlacement(hwnd)
        if placement[1] == win32con.SW_MAXIMIZE:  # 检查是否最大化
            # 将窗口恢复到正常状态
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)

        # 设置窗口的位置和尺寸
        win32gui.SetWindowPos(hwnd, win32con.HWND_TOP, x, y, w, h, win32con.SWP_SHOWWINDOW)
        # print(f"窗口 '{window_title}' 的位置和尺寸已设置为 ({x}, {y}, {w}, {h})")
        return True
    except Exception as e:
        # print(f"出错: {e}")
        return False


def mkdir(dir_path):
    if not os.path.exists(dir_path):  # 检查文件夹是否存在
        os.makedirs(dir_path)  # 递归创建文件夹
        return True
    else:
        return None


def make_screenshot(box, name, path):
    save_path = f"{path}/{name}.png"
    screenshot = pyautogui.screenshot(region=box)
    screenshot.save(save_path)
    logger.info(f"Created Screenshot, path: {save_path}")
    return None


def get_current_time(m=False):
    current_time = time.localtime()
    if m:
        return current_time.tm_mon
    formatted_time = time.strftime("%H:%M:%S", current_time)
    return formatted_time


def get_current_date(y=False, m=False, d=False, t=False):
    current_time = time.localtime()
    if y:
        formatted_time = time.strftime("%Y", current_time)
    elif m:
        formatted_time = time.strftime("%m", current_time)
    elif d:
        formatted_time = time.strftime("%d", current_time)
    elif t:
        formatted_time = time.strftime("%Y%m%d%H%M", current_time)
    else:
        formatted_time = time.strftime("%Y%m%d", current_time)
    return formatted_time


def pop_error_window(error_message, title='错误'):
    error_window = tk.Tk()  # 创建一个根窗口
    error_window.withdraw()  # 隐藏根窗口
    messagebox.showerror(title, error_message)  # 显示错误消息框
    error_window.quit()  # 退出主循环
    return None


def re_coord(coordinate):
    return int(coordinate * config.scale)


def click(x, y, box=position(0, 0, 0, 0)):  # 这个函数的坐标系是基于box的，不是整个屏幕的
    click_x = box.left + re_coord(random.randint(-3, 3) + x)
    click_y = box.top + re_coord(random.randint(-3, 3) + y)
    pyautogui.click(x=click_x, y=click_y)
    return None


def set_img_dir(step='shuiwu'):
    config.img_dir = f'./img/{step}/{int(config.scale * 100)}/'
    return config.img_dir


def set_scale():
    device = check_device()
    config.scale = device.scale_factor
    return None


def get_window_rect(window_title):
    # 查找窗口句柄
    hwnd = win32gui.FindWindow(None, window_title)
    if hwnd == 0:
        return False

    # 获取窗口的边界框
    rect = win32gui.GetWindowRect(hwnd)
    left, top, right, bottom = rect
    width = right - left
    height = bottom - top
    return left, top, width, height


def start_browser():
    edge_driver_path = 'msedgedriver.exe'  # 设置EdgeDriver路径
    service = Service(edge_driver_path)  # 创建Edge服务
    driver = webdriver.Edge(service=service)  # 创建Edge浏览器实例
    driver.set_window_size(1280, 768)  # 设置浏览器窗口大小
    driver.set_window_position(10, 10)  # 设置浏览器窗口位置
    return driver


def check_img(img_name, box, pos=False):
    try:
        img = pyautogui.locateOnScreen(f'{config.img_dir}{img_name}.png', region=box, confidence=0.9)
        if img is not None:
            if img_name != 'waiting_code_p1':
                logger.info(f"{img_name} Detected.")
            if pos:
                return img
            return True
        else:
            # logger.info(f"{img_name} Not Detected.")
            return False
    except pyautogui.ImageNotFoundException:
        # logger.info(f"{img_name} Not Detected.")
        return False
    except Exception as e:
        logger.error(f'Unknown Error Occurred While Checking Image {img_name}.\n')
        logger.exception(e)
        raise UnknownError(f"Unknown Error Occurred While Checking Image {img_name}.Info:{e}\n")


def check_404(box, step, display):  # 检测是否404，若404则尝试刷新，尝试三次后报错
    count = 0
    while True:
        try:
            refresh_box = (box.left + re_coord(206), box.top + re_coord(372),
                           box.width + re_coord(94), box.height + re_coord(48))
            page404 = pyautogui.locateOnScreen(f'{config.img_dir}refresh.png', region=refresh_box, confidence=0.9)
            if page404 is not None:
                logger.info(f'Refresh Button Detected.')
                if count == 3:
                    logger.error(f'Network Error In {step}.\n')
                    for i in range(5):
                        display.update(f"网络异常！页面加载失败！请检查网络设置！{5 - i} 秒后结束任务！", 'red')
                        time.sleep(1)
                    display.close()
                    raise NetworkError(f'Network Error In {step}.')
                pyautogui.press('f5')
                count += 1
                logger.warning(f"Page Loading Failed While {step}, Trying To Refresh For The {count}th Time.")
                display.update(f"页面加载失败，尝试第 {count} 次刷新", 'orange')
                time.sleep(5)
            else:
                return None
        except pyautogui.ImageNotFoundException:
            return None  # 若捕获ImageNotFoundException则跳出循环(未出现404)
        except Exception as e:
            # 捕获其他可能的异常
            logger.error(f'Unknown Error Occurred While Checking 404 In {step}.\n')
            logger.exception(e)
            for i in range(5):
                display.update(f"出现未知异常！{5 - i}秒后结束任务！", 'red')
                time.sleep(1)
            display.close()
            raise UnknownError(f'Unknown Error Occurred While Checking 404 In {step}.Info:{e}\n')


def box_center(box):
    x = int(box.left + box.width / 2)
    y = int(box.top + box.height / 2)
    return x, y


def check_device():
    # 设置 DPI 感知模式，不可或缺
    ctypes.windll.shcore.SetProcessDpiAwareness(1)

    # 获取物理分辨率
    hDC = win32gui.GetDC(0)
    phys_w = win32print.GetDeviceCaps(hDC, win32con.DESKTOPHORZRES)  # 横向物理分辨率
    phys_h = win32print.GetDeviceCaps(hDC, win32con.DESKTOPVERTRES)  # 纵向物理分辨率

    # 获取系统DPI以及缩放比例
    user32 = ctypes.WinDLL('user32')
    GetDpiForSystem = user32.GetDpiForSystem
    GetDpiForSystem.argtypes = [ctypes.c_uint32]
    GetDpiForSystem.restype = ctypes.c_uint32
    dpi = GetDpiForSystem(0)  # 缩放100%时DPI为96，125%时为120，150%时为144，175%时为168， 200%时为192
    scale_factor = dpi / 96.0  # 将DPI转换为缩放因子

    # 计算逻辑分辨率
    logic_w = int(phys_w / scale_factor)
    logic_h = int(phys_h / scale_factor)

    # 判断是否可以运行
    if logic_w < 1280 or logic_h < 768 or scale_factor >= 1.25:
        Available = False
    else:
        Available = True
    return deviceInfo(Available, logic_w, logic_h, phys_w, phys_h, scale_factor)


def extract_code(msg):
    # 使用正则表达式查找6位数的验证码
    match = re.search(r'\d{6}', msg)
    if match:
        return match.group(0)
    else:
        return False


def create_result_table(task):
    # 创建文件夹
    mkdir(config.result_dir)

    # 创建DataFrame
    data = {
        '序号': [''],
        '纳税人名称': [''],
        '纳税人类型': [''],
        f'{task}运行结果': ['']
    }
    df = pd.DataFrame(data)

    # 将DataFrame写入Excel文件
    file_path = f'{config.result_dir}{task}运行结果{get_current_date()}.xlsx'
    config.result_excel_path = file_path
    df.to_excel(file_path, index=False)

    # 加载Excel文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 设置City列的宽度
    ws.column_dimensions[get_column_letter(df.columns.get_loc('序号') + 1)].width = 8
    ws.column_dimensions[get_column_letter(df.columns.get_loc('纳税人名称') + 1)].width = 25
    ws.column_dimensions[get_column_letter(df.columns.get_loc('纳税人类型') + 1)].width = 13
    ws.column_dimensions[get_column_letter(df.columns.get_loc(f'{task}运行结果') + 1)].width = 80
    ws.row_dimensions[1].height = 15

    # 设置列对齐方式
    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws[get_column_letter(df.columns.get_loc('序号') + 1)]:
        row.alignment = alignment

    # 保存修改后的文件
    wb.save(file_path)
    return None


def write_result_table(serial, corp_name, corp_type, result, red=False):
    row_number = int(serial) + 1
    # 加载Excel文件
    wb = load_workbook(config.result_excel_path)
    ws = wb.active

    # 写入字段到指定行号
    ws.cell(row=row_number, column=1, value=serial)
    ws.cell(row=row_number, column=2, value=corp_name)
    ws.cell(row=row_number, column=3, value=corp_type)
    ws.cell(row=row_number, column=4, value=result)

    # 设置行高
    ws.row_dimensions[row_number].height = 40

    # 设置序号上下左右居中
    cell1 = ws.cell(row=row_number, column=1)
    cell1.alignment = Alignment(horizontal='center', vertical='center')

    # 设置其它字段上下居中
    cell2 = ws.cell(row=row_number, column=2)
    cell3 = ws.cell(row=row_number, column=3)
    cell4 = ws.cell(row=row_number, column=4)
    alignment_vertical_center = Alignment(vertical='center', wrap_text=True)
    cell2.alignment = alignment_vertical_center
    cell3.alignment = alignment_vertical_center
    cell4.alignment = alignment_vertical_center

    # 判断是否要标红第五列
    if red:
        red_font = Font(color="FF0000")  # 红色字体的RGB值
        ws.cell(row=row_number, column=4).font = red_font

    # 设置网格线（内外框线）
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 单元格设置边框
    for col in range(1, 5):
        cell = ws.cell(row=row_number, column=col)
        cell.border = thin_border

    # 保存文件
    wb.save(config.result_excel_path)
    return None


def insert_checkmark(row, col):
    # 加载文件
    excel_path = config.excel_path
    wb = load_workbook(excel_path)
    sheet = wb.active

    # 打勾
    sheet.cell(row=row, column=col).value = "√"

    # 设置上下左右居中
    cell = sheet.cell(row=row, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(excel_path)


def show_yn_popup(title, text, y_func=None, n_func=None, y_text='是', n_text='否'):
    # 标记是否选择
    has_chosen = False

    def yes_program():
        nonlocal has_chosen
        has_chosen = True
        root.destroy()
        if y_func is not None:
            y_func()

    def no_program():
        nonlocal has_chosen
        has_chosen = True
        root.destroy()
        if n_func is not None:
            n_func()

    def on_closing():
        if not has_chosen:
            sys.exit()  # 终止程序
        root.destroy()

    # 创建主窗口
    root = tk.Tk()
    root.title(title)
    root.geometry(f"{re_coord(260)}x{re_coord(120)}")
    root.attributes("-topmost", True)  # 设置窗口始终在最上层

    # 添加标签
    label = tk.Label(root, text=text)
    label.pack(pady=10)

    # 创建一个Frame放置按钮
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    # 添加yes按钮
    yes_button = tk.Button(button_frame, text=y_text, command=yes_program)
    yes_button.pack(side=tk.LEFT, padx=10)  # 左对齐并设置间距

    # 添加no按钮
    no_button = tk.Button(button_frame, text=n_text, command=no_program)
    no_button.pack(side=tk.LEFT, padx=10)  # 左对齐并设置间距

    # 监听窗口关闭事件
    root.protocol("WM_DELETE_WINDOW", on_closing)

    # 计算窗口位置并移动到屏幕中心
    root.update_idletasks()  # 更新窗口的几何信息
    screen_width = root.winfo_screenwidth()  # 获取屏幕宽度
    screen_height = root.winfo_screenheight()  # 获取屏幕高度
    window_width = root.winfo_width()  # 获取窗口宽度
    window_height = root.winfo_height()  # 获取窗口高度
    x = (screen_width // 2) - (window_width // 2)  # 计算 x 坐标
    y = (screen_height // 2) - (window_height // 2)  # 计算 y 坐标
    root.geometry(f"+{x}+{y}")  # 移动窗口到屏幕中心

    # 运行主循环
    root.mainloop()
    return None


class StatusDisplay:
    def __init__(self, text, position=(0, 0)):
        self.root = tk.Tk()
        self.root.overrideredirect(True)  # 去掉窗口的标题栏和边框
        self.root.geometry(f"+{position[0]}+{position[1]}")  # 设置窗口位置
        self.root.attributes("-topmost", True)  # 窗口置顶
        self.root.attributes("-alpha", 0.9)  # 设置窗口透明度
        self.root.attributes("-transparentcolor", "green")  # 设置白色为透明色
        self.label = tk.Label(self.root, text=text, bg="white", fg="black", font=("黑体", 13))
        self.label.pack()
        self.root.update()  # 在主线程中运行 tkinter 的主循环

    def update(self, text, color='black'):
        self.label.config(text=text, fg=color)
        self.root.update()

    def move_to(self, position):
        self.root.geometry(f"+{position[0]}+{position[1]}")
        self.root.update()

    def panel_style(self, left, top, width, height):
        self.label.config(bg="white", fg="red", font=("黑体", 13), anchor="w", justify="left")
        self.label.pack(fill="both", expand=True)  # Label 填充整个窗口
        self.root.geometry(f"{width}x{height}+{left}+{top}")  # 设置窗口位置和尺寸

    def close(self):
        self.root.destroy()


class LoginInfoError(Exception):
    pass


class CheckImgError(Exception):
    pass


class WeChatError(Exception):
    pass


class GetCodeError(Exception):
    pass


class NetworkError(Exception):
    pass


class ContactNotFount(Exception):
    pass


class UnknownError(Exception):
    pass


class TimeOutError(Exception):
    pass


class SkipError(Exception):
    pass


class RetryError(Exception):
    pass


if __name__ == '__main__':
    pop_error_window('此文件不可单独运行！！')
